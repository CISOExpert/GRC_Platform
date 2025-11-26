#!/usr/bin/env python3
"""
Import Threat Catalog and Risk Catalog from SCF Excel into database
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys

DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = "/Volumes/home/Projects/GRC_Unified_Platform/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx"

def clean_text(text):
    """Clean and normalize text"""
    if text is None:
        return None
    text = str(text).strip()
    while '\n\n' in text:
        text = text.replace('\n\n', '\n')
    return text if text else None

def extract_threats(excel_path):
    """Extract threats from Excel file"""
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Threat Catalog']
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"Threat Catalog columns: {headers[:5]}")
    
    threats = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        # Assuming structure: Threat Name, Category, Description
        threat_name = clean_text(row[0])
        if not threat_name:
            continue
            
        threat = {
            'name': threat_name,
            'description': clean_text(row[1]) if len(row) > 1 else None,
            'category': 'manmade',  # Default, adjust based on actual data
            'metadata': {}
        }
        
        # Store additional columns in metadata
        for i, header in enumerate(headers[2:10], start=2):
            if i < len(row) and row[i] and header:
                threat['metadata'][str(header).lower().replace(' ', '_')] = str(row[i])
        
        threats.append(threat)
    
    print(f"Extracted {len(threats)} threats")
    return threats

def extract_risks(excel_path):
    """Extract risk templates from Excel file"""
    print(f"Loading risk catalog...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Check if Risk Catalog sheet exists
    if 'Risk Catalog' not in wb.sheetnames:
        print("⚠️  Risk Catalog sheet not found")
        return []
    
    ws = wb['Risk Catalog']
    headers = [cell.value for cell in ws[1]]
    print(f"Risk Catalog columns: {headers[:5]}")
    
    risks = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        risk_name = clean_text(row[0])
        if not risk_name:
            continue
            
        risk = {
            'title': risk_name,
            'description': clean_text(row[1]) if len(row) > 1 else None,
            'category': 'operational',  # Default
            'metadata': {}
        }
        
        # Store additional columns in metadata
        for i, header in enumerate(headers[2:10], start=2):
            if i < len(row) and row[i] and header:
                risk['metadata'][str(header).lower().replace(' ', '_')] = str(row[i])
        
        risks.append(risk)
    
    print(f"Extracted {len(risks)} risk templates")
    return risks

def insert_threats(conn, threats):
    """Insert threats into database"""
    
    insert_sql = """
        INSERT INTO threats (name, category, description, threat_references, metadata)
        VALUES %s
        ON CONFLICT DO NOTHING
    """
    
    values = [
        (
            t['name'],
            t['category'],
            t['description'],
            None,  # threat_references
            psycopg2.extras.Json(t['metadata'])
        )
        for t in threats
    ]
    
    print(f"Inserting {len(values)} threats...")
    with conn.cursor() as cur:
        execute_values(cur, insert_sql, values)
        conn.commit()
    
    print("✅ Threats imported successfully!")

def create_risk_templates_table(conn):
    """Create risk templates table"""
    print("Creating risk templates table...")
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS risk_templates (
                id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
                title text NOT NULL,
                description text,
                category risk_category,
                suggested_ie numeric(5,2),
                suggested_ol numeric(5,2),
                metadata jsonb DEFAULT '{}'::jsonb,
                created_at timestamptz DEFAULT now()
            )
        """)
        
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_risk_templates_category 
            ON risk_templates(category)
        """)
        
        cur.execute("""
            COMMENT ON TABLE risk_templates IS 
            'Risk templates from SCF catalog for common risk scenarios'
        """)
        
        conn.commit()
    print("✅ Risk templates table ready")

def insert_risk_templates(conn, risks):
    """Insert risk templates into database"""
    
    if not risks:
        print("⚠️  No risk templates to insert")
        return
    
    insert_sql = """
        INSERT INTO risk_templates (title, description, category, metadata)
        VALUES %s
        ON CONFLICT DO NOTHING
    """
    
    values = [
        (
            r['title'],
            r['description'],
            r['category'],
            psycopg2.extras.Json(r['metadata'])
        )
        for r in risks
    ]
    
    print(f"Inserting {len(values)} risk templates...")
    with conn.cursor() as cur:
        execute_values(cur, insert_sql, values)
        conn.commit()
    
    print("✅ Risk templates imported successfully!")

def main():
    try:
        # Extract data
        threats = extract_threats(EXCEL_PATH)
        risks = extract_risks(EXCEL_PATH)
        
        if not threats and not risks:
            print("❌ No threat or risk data found")
            return 1
        
        print("\nConnecting to database...")
        conn = psycopg2.connect(DB_URL)
        
        # Import threats
        if threats:
            insert_threats(conn, threats)
        
        # Import risks
        if risks:
            create_risk_templates_table(conn)
            insert_risk_templates(conn, risks)
        
        # Verify
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM threats")
            threat_count = cur.fetchone()[0]
            print(f"\n✅ Import complete!")
            print(f"   Threats: {threat_count}")
            
            cur.execute("""
                SELECT COUNT(*) FROM information_schema.tables 
                WHERE table_name = 'risk_templates'
            """)
            if cur.fetchone()[0] > 0:
                cur.execute("SELECT COUNT(*) FROM risk_templates")
                risk_count = cur.fetchone()[0]
                print(f"   Risk Templates: {risk_count}")
            
            # Show sample threats
            cur.execute("""
                SELECT name, category, LEFT(description, 50) as desc
                FROM threats
                ORDER BY name
                LIMIT 5
            """)
            print("\nSample threats:")
            for row in cur.fetchall():
                print(f"  {row[0]} [{row[1]}]")
                if row[2]:
                    print(f"    {row[2]}...")
        
        conn.close()
        return 0
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
