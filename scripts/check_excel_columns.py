#!/usr/bin/env python3
import openpyxl

EXCEL_PATH = "/app/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb['SCF 2025.3.1']

# Check headers for risk/threat columns
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print('Looking for Risk/Threat columns...')
for i, h in enumerate(headers):
    if h and ('risk' in str(h).lower() or 'threat' in str(h).lower()):
        print(f'  Col {i}: {h}')

# Sample values from columns 296 and 336
print('\nColumn 296 (COL_RISK_SUMMARY) samples:')
samples = 0
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=50, values_only=True), start=2):
    val = row[296] if len(row) > 296 else None
    if val:
        print(f'  Row {row_num}: {str(val)[:100]}')
        samples += 1
    if samples >= 5:
        break

print('\nColumn 336 (COL_THREAT_SUMMARY) samples:')
samples = 0
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=50, values_only=True), start=2):
    val = row[336] if len(row) > 336 else None
    if val:
        print(f'  Row {row_num}: {str(val)[:100]}')
        samples += 1
    if samples >= 5:
        break
