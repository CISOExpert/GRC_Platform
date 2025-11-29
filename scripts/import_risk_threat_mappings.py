#!/usr/bin/env python3
"""
Import risk and threat control mappings from SCF Excel to database.
Uses external_controls for control references (unified model).
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys

DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = "/app/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx"

# Column indices
COL_CONTROL_ID = 2       # SCF #
COL_RISK_SUMMARY = 296   # Column KK - Risk Summary
COL_THREAT_SUMMARY = 336 # Column LY - Threat Summary

def clean_text(text):
    if text is None:
        return None
    text = str(text).strip()
    return text if text else None

def parse_id_list(text):
    """Parse newline or comma-separated list of IDs (e.g., 'R-AC-1\\nR-AC-2' or 'R-AC-1, R-AC-2')"""
    if not text:
        return []
    ids = []
    # First try newline split, then comma split
    text_str = str(text)
    if '\n' in text_str:
        items = text_str.split('\n')
    else:
        items = text_str.split(',')

    for item in items:
        item = item.strip()
        if item:
            ids.append(item)
    return ids

def import_risk_control_mappings(conn, ws):
    """Import risk-to-control mappings using external_controls"""
    print("\n" + "=" * 80)
    print("IMPORTING RISK-CONTROL MAPPINGS")
    print("=" * 80)

    cur = conn.cursor()

    # Get SCF framework ID
    cur.execute("SELECT id FROM frameworks WHERE code = 'SCF'")
    scf_framework_id = cur.fetchone()[0]

    # Get control mapping: ref_code -> external_controls.id (for SCF controls)
    cur.execute("""
        SELECT ref_code, id FROM external_controls
        WHERE framework_id = %s AND hierarchy_level = 'control'
    """, (scf_framework_id,))
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(control_map)} SCF controls in external_controls")

    # Get risk mapping: risk_id -> risks.id
    cur.execute("SELECT risk_id, id FROM risks WHERE risk_id IS NOT NULL")
    risk_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(risk_map)} risks")

    mappings = []
    controls_with_risks = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue

        control_uuid = control_map[control_id]

        # Parse risk summary column
        if len(row) > COL_RISK_SUMMARY:
            risk_summary = clean_text(row[COL_RISK_SUMMARY])
            risk_ids = parse_id_list(risk_summary)

            for risk_id in risk_ids:
                if risk_id in risk_map:
                    mappings.append((risk_map[risk_id], control_uuid, 'effective'))
                    controls_with_risks.add(control_id)

    print(f"Found {len(mappings)} risk-control mappings for {len(controls_with_risks)} controls")

    # Clear existing
    cur.execute("DELETE FROM risk_controls")

    # Insert new
    if mappings:
        execute_values(cur, """
            INSERT INTO risk_controls (risk_id, control_id, control_effectiveness)
            VALUES %s
            ON CONFLICT (risk_id, control_id) DO NOTHING
        """, mappings)

    conn.commit()
    print(f"✓ Imported {len(mappings)} risk-control mappings")

    # Update risk_count on external_controls
    cur.execute("""
        UPDATE external_controls ec
        SET risk_count = (
            SELECT COUNT(*) FROM risk_controls rc WHERE rc.control_id = ec.id
        )
        WHERE ec.framework_id = %s
    """, (scf_framework_id,))
    conn.commit()
    print("✓ Updated risk_count on external_controls")

    cur.close()
    return len(mappings)

def import_threat_control_mappings(conn, ws):
    """Import threat-to-control mappings using external_controls"""
    print("\n" + "=" * 80)
    print("IMPORTING THREAT-CONTROL MAPPINGS")
    print("=" * 80)

    cur = conn.cursor()

    # Get SCF framework ID
    cur.execute("SELECT id FROM frameworks WHERE code = 'SCF'")
    scf_framework_id = cur.fetchone()[0]

    # Get control mapping: ref_code -> external_controls.id (for SCF controls)
    cur.execute("""
        SELECT ref_code, id FROM external_controls
        WHERE framework_id = %s AND hierarchy_level = 'control'
    """, (scf_framework_id,))
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(control_map)} SCF controls in external_controls")

    # Get threat mapping: threat_id -> threats.id
    cur.execute("SELECT threat_id, id FROM threats WHERE threat_id IS NOT NULL")
    threat_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(threat_map)} threats")

    mappings = []
    controls_with_threats = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue

        control_uuid = control_map[control_id]

        # Parse threat summary column
        if len(row) > COL_THREAT_SUMMARY:
            threat_summary = clean_text(row[COL_THREAT_SUMMARY])
            threat_ids = parse_id_list(threat_summary)

            for threat_id in threat_ids:
                if threat_id in threat_map:
                    mappings.append((threat_map[threat_id], control_uuid, 'mitigates'))
                    controls_with_threats.add(control_id)

    print(f"Found {len(mappings)} threat-control mappings for {len(controls_with_threats)} controls")

    # Clear existing
    cur.execute("DELETE FROM threat_controls")

    # Insert new
    if mappings:
        execute_values(cur, """
            INSERT INTO threat_controls (threat_id, control_id, mitigation_level)
            VALUES %s
            ON CONFLICT (threat_id, control_id) DO NOTHING
        """, mappings)

    conn.commit()
    print(f"✓ Imported {len(mappings)} threat-control mappings")

    # Update threat_count on external_controls
    cur.execute("""
        UPDATE external_controls ec
        SET threat_count = (
            SELECT COUNT(*) FROM threat_controls tc WHERE tc.control_id = ec.id
        )
        WHERE ec.framework_id = %s
    """, (scf_framework_id,))
    conn.commit()
    print("✓ Updated threat_count on external_controls")

    cur.close()
    return len(mappings)

def main():
    print("=" * 80)
    print("RISK & THREAT CONTROL MAPPINGS IMPORT")
    print("=" * 80)

    print("\nLoading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['SCF 2025.3.1']

    print("Connecting to database...")
    conn = psycopg2.connect(DB_URL)

    # Import risk mappings
    risk_count = import_risk_control_mappings(conn, ws)

    # Reload worksheet for threat mappings
    ws = wb['SCF 2025.3.1']
    threat_count = import_threat_control_mappings(conn, ws)

    conn.close()

    print("\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    print(f"Risk-control mappings:    {risk_count}")
    print(f"Threat-control mappings:  {threat_count}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n✗ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
