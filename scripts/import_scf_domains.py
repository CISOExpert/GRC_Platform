#!/usr/bin/env python3
"""
Import SCF Domains & Principles from Excel into framework_metadata table.

This script reads the "SCF Domains & Principles" tab from the SCF Excel file
and imports the domain metadata (principles and intent) into the framework_metadata table.

Usage:
    python3 scripts/import_scf_domains.py

Requirements:
    - openpyxl
    - psycopg2-binary
"""

import json
import re
from pathlib import Path

try:
    import openpyxl
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install openpyxl psycopg2-binary")
    exit(1)


# Database connection settings (Supabase local dev)
DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 54322,
    "database": "postgres",
    "user": "postgres",
    "password": "postgres"
}

# Excel file path
EXCEL_FILE = Path(__file__).parent.parent / "reference_material" / "secure-controls-framework-scf-2025-3-1 (1).xlsx"
SHEET_NAME = "SCF Domains & Principles"


def clean_text(value):
    """Clean text value - handle None and strip whitespace."""
    if value is None:
        return None
    # Convert to string and clean
    text = str(value).strip()
    # Replace non-breaking spaces and other unicode whitespace
    text = re.sub(r'\s+', ' ', text)
    return text if text else None


def get_scf_framework_id(cursor):
    """Get the SCF framework ID from the database."""
    cursor.execute("SELECT id FROM frameworks WHERE code = 'SCF'")
    result = cursor.fetchone()
    if not result:
        raise ValueError("SCF framework not found in database. Run migrations first.")
    return result[0]


def read_domains_from_excel():
    """Read SCF Domains & Principles from Excel file."""
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    print(f"Reading from: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # Read headers
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    # Expected columns (by index)
    # 0: #
    # 1: SCF Domain
    # 2: SCF Identifier
    # 3: Cybersecurity & Data Privacy by Design (C|P) Principles
    # 4: Principle Intent

    domains = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[1] and not row[2]:
            continue

        domain_number = row[0]
        domain_name = clean_text(row[1])
        identifier = clean_text(row[2])
        principles = clean_text(row[3])
        principle_intent = clean_text(row[4])

        if not identifier:
            print(f"  Skipping row with no identifier: {row}")
            continue

        # Clean domain number (might have non-breaking space)
        if domain_number:
            domain_number = int(str(domain_number).strip().replace('\xa0', ''))

        domains.append({
            "domain_number": domain_number,
            "domain_name": domain_name,
            "identifier": identifier,
            "principles": principles,
            "principle_intent": principle_intent
        })

    wb.close()
    return domains


def import_domains(domains, framework_id, cursor):
    """Import domains into framework_metadata table."""
    print(f"\nImporting {len(domains)} domains...")

    # Prepare data for insertion
    records = []
    for domain in domains:
        metadata = {
            "type": "domain",
            "domain_number": domain["domain_number"],
            "domain_name": domain["domain_name"],
            "principles": domain["principles"],
            "principle_intent": domain["principle_intent"]
        }

        records.append((
            framework_id,
            domain["identifier"],  # reference_key (e.g., "GOV", "AAT")
            json.dumps(metadata)
        ))

    # Insert with ON CONFLICT to handle re-runs
    insert_sql = """
        INSERT INTO framework_metadata (framework_id, reference_key, metadata)
        VALUES %s
        ON CONFLICT (framework_id, reference_key)
        DO UPDATE SET
            metadata = EXCLUDED.metadata,
            updated_at = NOW()
    """

    execute_values(cursor, insert_sql, records)
    print(f"  Inserted/updated {len(records)} domain records")


def main():
    print("=" * 60)
    print("SCF Domains & Principles Import")
    print("=" * 60)

    # Read domains from Excel
    domains = read_domains_from_excel()
    print(f"\nFound {len(domains)} domains in Excel:")
    for d in domains[:5]:
        print(f"  {d['identifier']}: {d['domain_name'][:50]}...")
    if len(domains) > 5:
        print(f"  ... and {len(domains) - 5} more")

    # Connect to database
    print(f"\nConnecting to database at {DB_CONFIG['host']}:{DB_CONFIG['port']}...")
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    try:
        # Get SCF framework ID
        framework_id = get_scf_framework_id(cursor)
        print(f"SCF Framework ID: {framework_id}")

        # Import domains
        import_domains(domains, framework_id, cursor)

        # Commit
        conn.commit()
        print("\n✓ Import completed successfully!")

        # Verify
        cursor.execute("""
            SELECT reference_key, metadata->>'domain_name' as domain_name
            FROM framework_metadata
            WHERE framework_id = %s
            ORDER BY (metadata->>'domain_number')::int
            LIMIT 5
        """, (framework_id,))

        print("\nVerification (first 5 domains):")
        for row in cursor.fetchall():
            print(f"  {row[0]}: {row[1]}")

        cursor.execute("""
            SELECT COUNT(*) FROM framework_metadata WHERE framework_id = %s
        """, (framework_id,))
        total = cursor.fetchone()[0]
        print(f"\nTotal domain records: {total}")

    except Exception as e:
        conn.rollback()
        print(f"\n✗ Error: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    main()
