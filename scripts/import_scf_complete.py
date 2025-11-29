#!/usr/bin/env python3
"""
Complete SCF 2025.3.1 Data Import
Updates controls with SCRM tiers and errata, imports baselines, risk/threat mappings, and framework mappings
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys
import re

DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = "/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx"

# Column indices
COL_CONTROL_ID = 2
COL_SCRM_TIER1 = 19
COL_SCRM_TIER2 = 20
COL_SCRM_TIER3 = 21
COL_ERRATA = 374  # Column NO

# SCF CORE baseline columns (JZ-KG, indices 285-292)
BASELINE_COLUMNS = {
    285: 'community_derived',
    286: 'fundamentals',
    287: 'mad',
    288: 'esp_l1',
    289: 'esp_l2',
    290: 'esp_l3',
    291: 'ai_enabled',
    292: 'ai_model'
}

COL_RISK_SUMMARY = 296  # Column KK
COL_THREAT_SUMMARY = 336  # Column LY

# Framework mapping columns (AC-NJ, indices 28-269)
FRAMEWORK_MAPPING_START = 28
FRAMEWORK_MAPPING_END = 269

def clean_text(text):
    if text is None:
        return None
    text = str(text).strip()
    return text if text else None

def has_x(value):
    """Check if cell contains x or X"""
    if value is None:
        return False
    return 'x' in str(value).lower()

def parse_id_list(text):
    """Parse comma-separated list of IDs (e.g., 'R-AC-1, R-AC-2')"""
    if not text:
        return []
    ids = []
    for item in str(text).split(','):
        item = item.strip()
        if item:
            ids.append(item)
    return ids

def update_control_enhancements(conn, ws):
    """Update existing controls with SCRM tiers and errata"""
    print("\n" + "=" * 80)
    print("UPDATING CONTROL ENHANCEMENTS")
    print("=" * 80)
    
    cur = conn.cursor()
    updates = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id:
            continue
        
        scrm_tier1 = has_x(row[COL_SCRM_TIER1])
        scrm_tier2 = has_x(row[COL_SCRM_TIER2])
        scrm_tier3 = has_x(row[COL_SCRM_TIER3])
        errata = clean_text(row[COL_ERRATA]) if len(row) > COL_ERRATA else None
        
        updates.append((scrm_tier1, scrm_tier2, scrm_tier3, errata, control_id))
    
    print(f"Updating {len(updates)} controls...")
    
    cur.executemany("""
        UPDATE scf_controls
        SET scrm_tier1 = %s,
            scrm_tier2 = %s,
            scrm_tier3 = %s,
            errata_notes = %s
        WHERE control_id = %s
    """, updates)
    
    conn.commit()
    print(f"✓ Updated {len(updates)} controls with SCRM tiers and errata")
    cur.close()
    return len(updates)

def import_baselines(conn, ws):
    """Import SCF CORE baseline memberships"""
    print("\n" + "=" * 80)
    print("IMPORTING SCF CORE BASELINES")
    print("=" * 80)
    
    cur = conn.cursor()
    
    # Get control_id to UUID mapping
    cur.execute("SELECT control_id, id FROM scf_controls")
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    
    baseline_data = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue
        
        control_uuid = control_map[control_id]
        
        # Check each baseline column
        for col_idx, baseline_type in BASELINE_COLUMNS.items():
            if len(row) > col_idx:
                cell_value = clean_text(row[col_idx])
                if cell_value:  # Non-empty = included in this baseline
                    baseline_data.append((control_uuid, baseline_type))
    
    print(f"Inserting {len(baseline_data)} baseline memberships...")
    
    # Clear existing baselines
    cur.execute("DELETE FROM control_baselines")
    
    # Insert new baseline data
    execute_values(cur, """
        INSERT INTO control_baselines (control_id, baseline_type, included)
        VALUES %s
        ON CONFLICT (control_id, baseline_type) DO NOTHING
    """, [(c, b, True) for c, b in baseline_data])
    
    conn.commit()
    print(f"✓ Imported {len(baseline_data)} baseline memberships across 8 baseline types")
    cur.close()
    return len(baseline_data)

def import_risk_control_mappings(conn, ws):
    """Import risk-to-control mappings"""
    print("\n" + "=" * 80)
    print("IMPORTING RISK-CONTROL MAPPINGS")
    print("=" * 80)
    
    cur = conn.cursor()
    
    # Get mappings
    cur.execute("SELECT control_id, id FROM scf_controls")
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    
    cur.execute("SELECT risk_id, id FROM risks WHERE risk_id IS NOT NULL")
    risk_map = {row[0]: row[1] for row in cur.fetchall()}
    
    mappings = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue
        
        control_uuid = control_map[control_id]
        
        # Parse risk summary column
        if len(row) > COL_RISK_SUMMARY:
            risk_summary = clean_text(row[COL_RISK_SUMMARY])
            risk_ids = parse_id_list(risk_summary)
            
            for risk_id in risk_ids:
                if risk_id in risk_map:
                    mappings.append((risk_map[risk_id], control_uuid, 'effective'))
    
    print(f"Inserting {len(mappings)} risk-control mappings...")
    
    # Clear existing
    cur.execute("DELETE FROM risk_controls")
    
    # Insert new
    execute_values(cur, """
        INSERT INTO risk_controls (risk_id, control_id, control_effectiveness)
        VALUES %s
        ON CONFLICT (risk_id, control_id) DO NOTHING
    """, mappings)
    
    conn.commit()
    print(f"✓ Imported {len(mappings)} risk-control mappings")
    cur.close()
    return len(mappings)

def import_threat_control_mappings(conn, ws):
    """Import threat-to-control mappings"""
    print("\n" + "=" * 80)
    print("IMPORTING THREAT-CONTROL MAPPINGS")
    print("=" * 80)
    
    cur = conn.cursor()
    
    # Get mappings
    cur.execute("SELECT control_id, id FROM scf_controls")
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    
    cur.execute("SELECT threat_id, id FROM threats WHERE threat_id IS NOT NULL")
    threat_map = {row[0]: row[1] for row in cur.fetchall()}
    
    mappings = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue
        
        control_uuid = control_map[control_id]
        
        # Parse threat summary column
        if len(row) > COL_THREAT_SUMMARY:
            threat_summary = clean_text(row[COL_THREAT_SUMMARY])
            threat_ids = parse_id_list(threat_summary)
            
            for threat_id in threat_ids:
                if threat_id in threat_map:
                    mappings.append((threat_map[threat_id], control_uuid, 'mitigates'))
    
    print(f"Inserting {len(mappings)} threat-control mappings...")
    
    # Clear existing
    cur.execute("DELETE FROM threat_controls")
    
    # Insert new
    execute_values(cur, """
        INSERT INTO threat_controls (threat_id, control_id, mitigation_level)
        VALUES %s
        ON CONFLICT (threat_id, control_id) DO NOTHING
    """, mappings)
    
    conn.commit()
    print(f"✓ Imported {len(mappings)} threat-control mappings")
    cur.close()
    return len(mappings)

def import_framework_mappings(conn, ws, headers):
    """Import framework mappings from columns AC-NJ"""
    print("\n" + "=" * 80)
    print("IMPORTING FRAMEWORK MAPPINGS")
    print("=" * 80)
    
    cur = conn.cursor()
    
    # Get control mapping
    cur.execute("SELECT control_id, id FROM scf_controls")
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    
    # Get framework mapping by column header
    cur.execute("SELECT id, mapping_column_header FROM frameworks WHERE mapping_column_header IS NOT NULL")
    framework_by_header = {clean_text(row[1]): row[0] for row in cur.fetchall()}
    
    print(f"Found {len(framework_by_header)} frameworks with column headers")
    
    # First pass: collect all unique external controls and create them
    print("Pass 1: Creating external controls...")
    external_controls_to_create = {}  # (framework_id, ref_code) -> description
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue
        
        # Check each framework column
        for col_idx in range(FRAMEWORK_MAPPING_START, min(FRAMEWORK_MAPPING_END + 1, len(row))):
            cell_value = clean_text(row[col_idx])
            if not cell_value:
                continue
            
            # Get column header
            if col_idx < len(headers):
                header = clean_text(headers[col_idx])
                if header in framework_by_header:
                    framework_id = framework_by_header[header]
                    key = (framework_id, cell_value)
                    if key not in external_controls_to_create:
                        external_controls_to_create[key] = cell_value  # Use ref_code as description if no better option
    
    print(f"Creating {len(external_controls_to_create)} external controls...")
    
    # Create external controls
    for (framework_id, ref_code), description in external_controls_to_create.items():
        cur.execute("""
            INSERT INTO external_controls (framework_id, ref_code, description)
            VALUES (%s, %s, %s)
            ON CONFLICT (framework_id, ref_code) DO NOTHING
        """, (framework_id, ref_code, description))
    
    conn.commit()
    
    # Second pass: get external control UUIDs and create mappings
    print("Pass 2: Creating control mappings...")
    cur.execute("SELECT framework_id, ref_code, id FROM external_controls")
    external_control_map = {(row[0], row[1]): row[2] for row in cur.fetchall()}
    
    mappings = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id or control_id not in control_map:
            continue
        
        control_uuid = control_map[control_id]
        
        # Check each framework column
        for col_idx in range(FRAMEWORK_MAPPING_START, min(FRAMEWORK_MAPPING_END + 1, len(row))):
            cell_value = clean_text(row[col_idx])
            if not cell_value:
                continue
            
            # Get column header
            if col_idx < len(headers):
                header = clean_text(headers[col_idx])
                if header in framework_by_header:
                    framework_id = framework_by_header[header]
                    key = (framework_id, cell_value)
                    if key in external_control_map:
                        external_control_id = external_control_map[key]
                        mappings.append((control_uuid, external_control_id, framework_id))
    
    print(f"Inserting {len(mappings)} framework mappings...")
    
    # Clear existing mappings
    cur.execute("DELETE FROM scf_control_mappings")
    
    # Insert new mappings
    execute_values(cur, """
        INSERT INTO scf_control_mappings (scf_control_id, external_control_id, framework_id)
        VALUES %s
        ON CONFLICT (scf_control_id, external_control_id) DO NOTHING
    """, mappings)
    
    conn.commit()
    print(f"✓ Imported {len(external_controls_to_create)} external controls and {len(mappings)} mappings")
    cur.close()
    return len(mappings)

def main():
    print("=" * 80)
    print("SCF 2025.3.1 COMPLETE DATA IMPORT")
    print("=" * 80)
    
    print("\nLoading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['SCF 2025.3.1']
    
    # Get headers for framework mapping
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    
    print("Connecting to database...")
    conn = psycopg2.connect(DB_URL)
    
    # Run imports
    control_updates = update_control_enhancements(conn, ws)
    
    # Reload worksheet for baseline import
    ws = wb['SCF 2025.3.1']
    baseline_count = import_baselines(conn, ws)
    
    # Reload for risk mappings
    ws = wb['SCF 2025.3.1']
    risk_mapping_count = import_risk_control_mappings(conn, ws)
    
    # Reload for threat mappings
    ws = wb['SCF 2025.3.1']
    threat_mapping_count = import_threat_control_mappings(conn, ws)
    
    # Reload for framework mappings
    ws = wb['SCF 2025.3.1']
    framework_mapping_count = import_framework_mappings(conn, ws, headers)
    
    conn.close()
    
    print("\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    print(f"Controls updated:            {control_updates}")
    print(f"Baseline memberships:        {baseline_count}")
    print(f"Risk-control mappings:       {risk_mapping_count}")
    print(f"Threat-control mappings:     {threat_mapping_count}")
    print(f"Framework mappings:          {framework_mapping_count}")
    print("\n✓ Complete SCF data import successful!")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n✗ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
