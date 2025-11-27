import os
from typing import Any, Dict, List, Tuple

import openpyxl
import psycopg2


EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "reference_material", "secure-controls-framework-scf-2025-3-1 (1).xlsx")
SCF_SHEET_NAME = "SCF 2025.3.1"


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    value = value.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    value = " ".join(value.split())
    return value.strip()


def get_db_connection():
    return psycopg2.connect(
        dbname=os.getenv("SUPABASE_DB_NAME", "postgres"),
        user=os.getenv("SUPABASE_DB_USER", "postgres"),
        password=os.getenv("SUPABASE_DB_PASSWORD", "postgres"),
        host=os.getenv("SUPABASE_DB_HOST", "127.0.0.1"),
        port=int(os.getenv("SUPABASE_DB_PORT", "54322")),
    )


def load_scf_controls(conn) -> Dict[int, str]:
    """Return mapping of SCF Excel row index -> scf_controls.id via SCF # (column C).

    Assumes the SCF sheet's column C (3) holds the SCF # values
    like IRO-12.3, NET-04.8, etc., which match scf_controls.control_id.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT id, control_id FROM scf_controls;")
        rows = cur.fetchall()
    by_control_id = {control_id: scf_id for scf_id, control_id in rows}

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = wb[SCF_SHEET_NAME]
    mapping: Dict[int, str] = {}
    for row_idx in range(2, sheet.max_row + 1):
        control_cell = sheet.cell(row=row_idx, column=3)  # column C: SCF #
        control_id = (control_cell.value or "").strip() if isinstance(control_cell.value, str) else control_cell.value
        if not control_id:
            continue
        scf_id = by_control_id.get(control_id)
        if scf_id:
            mapping[row_idx] = scf_id
    wb.close()
    return mapping


def load_frameworks_needing_mappings(conn) -> List[Tuple[str, str, int]]:
    """Return list of (framework_id, mapping_column_header, column_index) for frameworks with 0 mappings.

    Column index is determined by matching normalized mapping_column_header against Excel headers.
    """
    # First, collect Excel headers by normalized text -> column index
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = wb[SCF_SHEET_NAME]
    excel_headers: Dict[str, int] = {}
    for col_idx in range(29, 274):
        cell = sheet.cell(row=1, column=col_idx)
        norm = normalize_header(cell.value)
        if norm and norm not in excel_headers:
            excel_headers[norm] = col_idx
    wb.close()

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT f.id, f.mapping_column_header, fc.mapping_count
            FROM frameworks f
            JOIN get_frameworks_with_counts() fc ON fc.code = f.code
            WHERE f.mapping_column_header IS NOT NULL
            ORDER BY fc.mapping_count
            """
        )
        rows = cur.fetchall()

    result: List[Tuple[str, str, int]] = []
    for framework_id, header, mapping_count in rows:
        if mapping_count > 0:
            continue
        norm_header = normalize_header(header)
        col_idx = excel_headers.get(norm_header)
        if not col_idx:
            # No corresponding Excel column; skip
            continue
        result.append((framework_id, norm_header, col_idx))
    return result


def ensure_external_control(conn, framework_id: str, ref_code: str) -> str:
    """Get or create an external_controls row and return its id."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id FROM external_controls
            WHERE framework_id = %s AND ref_code = %s
            """,
            (framework_id, ref_code),
        )
        row = cur.fetchone()
        if row:
            return row[0]

        cur.execute(
            """
            INSERT INTO external_controls (framework_id, ref_code, description)
            VALUES (%s, %s, %s)
            RETURNING id
            """,
            (framework_id, ref_code, f"External control {ref_code}"),
        )
        new_id = cur.fetchone()[0]
        return new_id


def import_mappings_for_framework(conn, framework_id: str, header: str, col_idx: int, scf_row_to_id: Dict[int, str]) -> int:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = wb[SCF_SHEET_NAME]

    created_mappings = 0
    with conn.cursor() as cur:
        for row_idx, scf_id in scf_row_to_id.items():
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if not value:
                continue
            if isinstance(value, str):
                ref = value.strip()
            else:
                ref = str(value).strip()
            if not ref:
                continue

            external_id = ensure_external_control(conn, framework_id, ref)

            cur.execute(
                """
                INSERT INTO scf_control_mappings (scf_control_id, external_control_id, framework_id)
                VALUES (%s, %s, %s)
                ON CONFLICT (scf_control_id, external_control_id, framework_id) DO NOTHING
                """,
                (scf_id, external_id, framework_id),
            )
            if cur.rowcount > 0:
                created_mappings += 1

    wb.close()
    return created_mappings


def main() -> None:
    print("Connecting to database...")
    conn = get_db_connection()
    conn.autocommit = False
    try:
        print("Loading SCF control row mapping (Excel row -> scf_controls.id)...")
        scf_row_to_id = load_scf_controls(conn)
        print(f"Mapped {len(scf_row_to_id)} SCF Excel rows to scf_controls records.")

        print("Finding frameworks with Excel columns but zero mappings...")
        frameworks = load_frameworks_needing_mappings(conn)
        print(f"Frameworks with zero mappings and a matching Excel column: {len(frameworks)}")

        total_created = 0
        for framework_id, header, col_idx in frameworks:
            print(f"\nImporting mappings for framework header '{header}' (column {col_idx})...")
            created = import_mappings_for_framework(conn, framework_id, header, col_idx, scf_row_to_id)
            print(f"Created {created} new scf_control_mappings rows for '{header}'.")
            total_created += created

        conn.commit()
        print(f"\nDone. Total new mappings created: {total_created}")
    except Exception as exc:
        print("Error during import, rolling back:", exc)
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
