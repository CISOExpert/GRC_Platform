#!/usr/bin/env python3
"""
Complete fix: Analyze headers, update database, and re-import all mappings
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys
import re

DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = "/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx"

COL_CONTROL_ID = 2
FRAMEWORK_MAPPING_START = 28
FRAMEWORK_MAPPING_END = 269

def normalize_for_comparison(text):
    """Normalize text for comparison"""
    if text is None:
        return None
    text = str(text).strip()
    normalized = re.sub(r'\s+', ' ', text)
    return normalized if normalized else None

def clean_text(text):
    if text is None:
        return None
    text = str(text).strip()
    return text if text else None

print("=" * 80)
print("COMPLETE FRAMEWORK MAPPING FIX")
print("=" * 80)

# Load Excel once
print("\nLoading Excel workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb['SCF 2025.3.1']
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

# Connect to database
print("Connecting to database...")
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# ============================================================================
# STEP 1: FIX FRAMEWORK HEADERS
# ============================================================================
print("\n" + "=" * 80)
print("STEP 1: FIXING FRAMEWORK HEADERS")
print("=" * 80)

# Get Excel headers with newlines
excel_headers = {}
for col_idx in range(FRAMEWORK_MAPPING_START, min(FRAMEWORK_MAPPING_END + 1, len(headers))):
    if headers[col_idx]:
        original = str(headers[col_idx]).strip()
        if original:
            excel_headers[col_idx] = original

print(f"Found {len(excel_headers)} Excel framework headers")

# Get database frameworks
cur.execute("""
    SELECT id, code, version, mapping_column_header
    FROM frameworks
    WHERE mapping_column_header IS NOT NULL
    ORDER BY code, version
""")
db_frameworks = cur.fetchall()
print(f"Found {len(db_frameworks)} database frameworks with headers")

# Build normalized lookup
excel_normalized_to_original = {}
for idx, header in excel_headers.items():
    normalized = normalize_for_comparison(header)
    if normalized:
        excel_normalized_to_original[normalized] = header

# Compare and fix
mismatches = []
for fw_id, code, version, db_header in db_frameworks:
    db_normalized = normalize_for_comparison(db_header)
    exact_match = db_header in excel_headers.values()
    
    if not exact_match and db_normalized in excel_normalized_to_original:
        excel_header = excel_normalized_to_original[db_normalized]
        mismatches.append((excel_header, code, version))

print(f"\nFound {len(mismatches)} headers to fix")

if mismatches:
    print("Applying fixes...")
    for excel_header, code, version in mismatches:
        if version:
            cur.execute(
                "UPDATE frameworks SET mapping_column_header = %s WHERE code = %s AND version = %s",
                (excel_header, code, version)
            )
        else:
            cur.execute(
                "UPDATE frameworks SET mapping_column_header = %s WHERE code = %s AND version IS NULL",
                (excel_header, code)
            )
    conn.commit()
    print(f"✓ Fixed {len(mismatches)} framework headers")

# ============================================================================
# STEP 2: RE-IMPORT FRAMEWORK MAPPINGS
# ============================================================================
print("\n" + "=" * 80)
print("STEP 2: RE-IMPORTING FRAMEWORK MAPPINGS")
print("=" * 80)

# Reload worksheet
ws = wb['SCF 2025.3.1']
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

# Get control mapping
cur.execute("SELECT control_id, id FROM scf_controls")
control_map = {row[0]: row[1] for row in cur.fetchall()}
print(f"Loaded {len(control_map)} SCF controls")

# Get framework mapping (now with corrected headers)
cur.execute("SELECT id, mapping_column_header FROM frameworks WHERE mapping_column_header IS NOT NULL")
framework_by_header = {row[1]: row[0] for row in cur.fetchall()}
print(f"Loaded {len(framework_by_header)} frameworks")

# Pass 1: Create external controls
print("\nPass 1: Creating external controls...")
external_controls_to_create = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    control_id = clean_text(row[COL_CONTROL_ID])
    if not control_id or control_id not in control_map:
        continue
    
    for col_idx in range(FRAMEWORK_MAPPING_START, min(FRAMEWORK_MAPPING_END + 1, len(row))):
        cell_value = clean_text(row[col_idx])
        if not cell_value:
            continue
        
        if col_idx < len(headers):
            header = headers[col_idx]
            if header in framework_by_header:
                framework_id = framework_by_header[header]
                key = (framework_id, cell_value)
                if key not in external_controls_to_create:
                    external_controls_to_create[key] = cell_value

print(f"Creating {len(external_controls_to_create)} external controls...")

for (framework_id, ref_code), description in external_controls_to_create.items():
    cur.execute("""
        INSERT INTO external_controls (framework_id, ref_code, description)
        VALUES (%s, %s, %s)
        ON CONFLICT (framework_id, ref_code) DO NOTHING
    """, (framework_id, ref_code, description))

conn.commit()

# Pass 2: Create mappings
print("Pass 2: Creating control mappings...")
ws = wb['SCF 2025.3.1']
cur.execute("SELECT framework_id, ref_code, id FROM external_controls")
external_control_map = {(row[0], row[1]): row[2] for row in cur.fetchall()}

mappings = []

for row in ws.iter_rows(min_row=2, values_only=True):
    control_id = clean_text(row[COL_CONTROL_ID])
    if not control_id or control_id not in control_map:
        continue
    
    control_uuid = control_map[control_id]
    
    for col_idx in range(FRAMEWORK_MAPPING_START, min(FRAMEWORK_MAPPING_END + 1, len(row))):
        cell_value = clean_text(row[col_idx])
        if not cell_value:
            continue
        
        if col_idx < len(headers):
            header = headers[col_idx]
            if header in framework_by_header:
                framework_id = framework_by_header[header]
                key = (framework_id, cell_value)
                if key in external_control_map:
                    external_control_id = external_control_map[key]
                    mappings.append((control_uuid, external_control_id, framework_id))

print(f"Inserting {len(mappings)} mappings...")

cur.execute("DELETE FROM scf_control_mappings")
execute_values(cur, """
    INSERT INTO scf_control_mappings (scf_control_id, external_control_id, framework_id)
    VALUES %s
    ON CONFLICT (scf_control_id, external_control_id) DO NOTHING
""", mappings)

conn.commit()

# ============================================================================
# STEP 3: VERIFY RESULTS
# ============================================================================
print("\n" + "=" * 80)
print("STEP 3: VERIFICATION - TOP 25 FRAMEWORKS")
print("=" * 80)

cur.execute("""
    SELECT 
      f.code, 
      f.version,
      COUNT(DISTINCT scm.id) as mapping_count,
      COUNT(DISTINCT scm.scf_control_id) as scf_controls
    FROM frameworks f
    LEFT JOIN external_controls ec ON ec.framework_id = f.id
    LEFT JOIN scf_control_mappings scm ON scm.external_control_id = ec.id
    WHERE scm.id IS NOT NULL
    GROUP BY f.id, f.code, f.version
    ORDER BY mapping_count DESC
    LIMIT 25;
""")

print(f"\n{'Code':<35} {'Version':<15} {'Mappings':>10} {'SCF Controls':>13}")
print("-" * 80)
for row in cur.fetchall():
    code, version, mapping_count, scf_count = row
    version_str = version if version else "N/A"
    print(f"{code:<35} {version_str:<15} {mapping_count:>10} {scf_count:>13}")

cur.close()
conn.close()
wb.close()

print("\n" + "=" * 80)
print("COMPLETE! ALL FRAMEWORK MAPPINGS FIXED AND RE-IMPORTED")
print("=" * 80)
print(f"✓ Fixed {len(mismatches)} framework headers")
print(f"✓ Created {len(external_controls_to_create)} external controls")
print(f"✓ Imported {len(mappings)} control mappings")
