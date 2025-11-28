#!/usr/bin/env python3
"""
Complete Threat and Risk Import for GRC Platform
Imports:
1. Threat Catalog → threats table
2. Risk Catalog → risks table
3. Control→Threat mappings → threat_controls table
4. Control→Risk mappings → risk_controls table

Updated 2025-11-28 to use external_controls (unified table) instead of deprecated scf_controls
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys
import re
import os

# Configuration
DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "reference_material",
    "secure-controls-framework-scf-2025-3-1 (1).xlsx"
)

# Column indices for SCF 2025.3.1 tab (0-indexed)
COL_CONTROL_ID = 2  # Column C
COL_RISK_SUMMARY = 296  # Column KK - Risk Threat Summary
COL_THREAT_SUMMARY = 336  # Column LY - Control Threat Summary

def clean_text(text):
    """Clean and normalize text"""
    if text is None:
        return None
    text = str(text).strip()
    text = re.sub(r'\n+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text if text else None

def parse_id_list(text):
    """Parse space or comma-separated list of IDs (e.g., 'R-AC-1 R-AC-2' or 'R-AC-1, R-AC-2')"""
    if not text:
        return []
    ids = []
    # Split on both spaces and commas
    text = str(text).replace(',', ' ')
    for item in text.split():
        item = item.strip()
        if item:
            ids.append(item)
    return ids

def import_threat_catalog(conn, wb):
    """Import threats from Threat Catalog tab"""
    print("\n" + "=" * 80)
    print("IMPORTING THREAT CATALOG")
    print("=" * 80)

    cur = conn.cursor()
    ws = wb['Threat Catalog']

    threats = []
    current_grouping = None

    # Start from row 8 (after headers based on original script)
    for row in ws.iter_rows(min_row=8, values_only=True):
        grouping = clean_text(row[0])
        threat_id = clean_text(row[1])
        name = clean_text(row[2])
        description = clean_text(row[3])

        # Track current grouping
        if grouping:
            current_grouping = grouping

        # Only process rows with threat IDs (NT-*, MT-*, FT-*)
        if threat_id and (threat_id.startswith('NT-') or threat_id.startswith('MT-') or threat_id.startswith('FT-')):
            threats.append({
                'threat_id': threat_id,
                'name': name,
                'threat_grouping': current_grouping,
                'description': description,
                'category': 'natural' if threat_id.startswith('NT-') else 'manmade'
            })

    print(f"Extracted {len(threats)} threats from Excel")

    # Insert threats
    inserted = 0
    for threat in threats:
        try:
            cur.execute("""
                INSERT INTO threats (threat_id, name, threat_grouping, description, category)
                VALUES (%s, %s, %s, %s, %s::threat_category)
                ON CONFLICT (threat_id) DO UPDATE
                SET name = EXCLUDED.name,
                    threat_grouping = EXCLUDED.threat_grouping,
                    description = EXCLUDED.description,
                    category = EXCLUDED.category
            """, (
                threat['threat_id'],
                threat['name'],
                threat['threat_grouping'],
                threat['description'],
                threat['category']
            ))
            inserted += 1
        except Exception as e:
            print(f"  Warning: Failed to insert threat {threat['threat_id']}: {e}")

    conn.commit()
    print(f"✓ Imported {inserted} threats")
    cur.close()
    return inserted

def import_risk_catalog(conn, wb):
    """Import risks from Risk Catalog tab"""
    print("\n" + "=" * 80)
    print("IMPORTING RISK CATALOG")
    print("=" * 80)

    cur = conn.cursor()
    ws = wb['Risk Catalog']

    risks = []
    current_grouping = None

    # Start from row 8 (after headers)
    for row in ws.iter_rows(min_row=8, values_only=True):
        grouping = clean_text(row[0])
        risk_id = clean_text(row[1])
        name = clean_text(row[2])
        description = clean_text(row[3])
        nist_csf_function = clean_text(row[4]) if len(row) > 4 else None

        # Track current grouping
        if grouping:
            current_grouping = grouping

        # Only process rows with risk IDs (R-*)
        if risk_id and risk_id.startswith('R-'):
            risks.append({
                'risk_id': risk_id,
                'title': name,
                'risk_grouping': current_grouping,
                'description': description,
                'nist_csf_function': nist_csf_function
            })

    print(f"Extracted {len(risks)} risks from Excel")

    # Insert risks (catalog risks have no org_id)
    inserted = 0
    for risk in risks:
        try:
            cur.execute("""
                INSERT INTO risks (risk_id, title, risk_grouping, description, nist_csf_function, status)
                VALUES (%s, %s, %s, %s, %s, 'catalog')
                ON CONFLICT (risk_id) DO UPDATE
                SET title = EXCLUDED.title,
                    risk_grouping = EXCLUDED.risk_grouping,
                    description = EXCLUDED.description,
                    nist_csf_function = EXCLUDED.nist_csf_function
            """, (
                risk['risk_id'],
                risk['title'],
                risk['risk_grouping'],
                risk['description'],
                risk['nist_csf_function']
            ))
            inserted += 1
        except Exception as e:
            print(f"  Warning: Failed to insert risk {risk['risk_id']}: {e}")

    conn.commit()
    print(f"✓ Imported {inserted} risks")
    cur.close()
    return inserted

def import_risk_control_mappings(conn, wb):
    """Import risk-to-control mappings from SCF 2025.3.1 tab"""
    print("\n" + "=" * 80)
    print("IMPORTING RISK-CONTROL MAPPINGS")
    print("=" * 80)

    cur = conn.cursor()
    ws = wb['SCF 2025.3.1']

    # Get control_id to UUID mapping from external_controls (SCF framework)
    cur.execute("""
        SELECT ec.ref_code, ec.id
        FROM external_controls ec
        JOIN frameworks f ON ec.framework_id = f.id
        WHERE f.code = 'SCF'
    """)
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(control_map)} SCF controls in external_controls")

    # Get risk_id to UUID mapping
    cur.execute("SELECT risk_id, id FROM risks WHERE risk_id IS NOT NULL")
    risk_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(risk_map)} risks in database")

    mappings = []
    missing_controls = set()
    missing_risks = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id:
            continue

        if control_id not in control_map:
            missing_controls.add(control_id)
            continue

        control_uuid = control_map[control_id]

        # Parse risk summary column (KK)
        if len(row) > COL_RISK_SUMMARY:
            risk_summary = clean_text(row[COL_RISK_SUMMARY])
            risk_ids = parse_id_list(risk_summary)

            for risk_id in risk_ids:
                if risk_id in risk_map:
                    mappings.append((risk_map[risk_id], control_uuid, 'effective'))
                else:
                    missing_risks.add(risk_id)

    if missing_controls:
        print(f"  Warning: {len(missing_controls)} control IDs not found in external_controls")
    if missing_risks:
        print(f"  Warning: {len(missing_risks)} risk IDs not found in risks table")
        print(f"    Sample missing risks: {list(missing_risks)[:5]}")

    print(f"Inserting {len(mappings)} risk-control mappings...")

    # Clear existing mappings
    cur.execute("DELETE FROM risk_controls")

    # Insert new mappings
    if mappings:
        execute_values(cur, """
            INSERT INTO risk_controls (risk_id, control_id, control_effectiveness)
            VALUES %s
            ON CONFLICT (risk_id, control_id) DO NOTHING
        """, mappings)

    conn.commit()
    print(f"✓ Imported {len(mappings)} risk-control mappings")
    cur.close()
    return len(mappings)

def import_threat_control_mappings(conn, wb):
    """Import threat-to-control mappings from SCF 2025.3.1 tab"""
    print("\n" + "=" * 80)
    print("IMPORTING THREAT-CONTROL MAPPINGS")
    print("=" * 80)

    cur = conn.cursor()
    ws = wb['SCF 2025.3.1']

    # Get control_id to UUID mapping from external_controls (SCF framework)
    cur.execute("""
        SELECT ec.ref_code, ec.id
        FROM external_controls ec
        JOIN frameworks f ON ec.framework_id = f.id
        WHERE f.code = 'SCF'
    """)
    control_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(control_map)} SCF controls in external_controls")

    # Get threat_id to UUID mapping
    cur.execute("SELECT threat_id, id FROM threats WHERE threat_id IS NOT NULL")
    threat_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Found {len(threat_map)} threats in database")

    mappings = []
    missing_controls = set()
    missing_threats = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        control_id = clean_text(row[COL_CONTROL_ID])
        if not control_id:
            continue

        if control_id not in control_map:
            missing_controls.add(control_id)
            continue

        control_uuid = control_map[control_id]

        # Parse threat summary column (LY)
        if len(row) > COL_THREAT_SUMMARY:
            threat_summary = clean_text(row[COL_THREAT_SUMMARY])
            threat_ids = parse_id_list(threat_summary)

            for threat_id in threat_ids:
                if threat_id in threat_map:
                    mappings.append((threat_map[threat_id], control_uuid, 'mitigates'))
                else:
                    missing_threats.add(threat_id)

    if missing_controls:
        print(f"  Warning: {len(missing_controls)} control IDs not found in external_controls")
    if missing_threats:
        print(f"  Warning: {len(missing_threats)} threat IDs not found in threats table")
        print(f"    Sample missing threats: {list(missing_threats)[:5]}")

    print(f"Inserting {len(mappings)} threat-control mappings...")

    # Clear existing mappings
    cur.execute("DELETE FROM threat_controls")

    # Insert new mappings
    if mappings:
        execute_values(cur, """
            INSERT INTO threat_controls (threat_id, control_id, mitigation_level)
            VALUES %s
            ON CONFLICT (threat_id, control_id) DO NOTHING
        """, mappings)

    conn.commit()
    print(f"✓ Imported {len(mappings)} threat-control mappings")
    cur.close()
    return len(mappings)

def verify_import(conn):
    """Verify import results"""
    print("\n" + "=" * 80)
    print("VERIFICATION")
    print("=" * 80)

    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM threats")
    threat_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM risks")
    risk_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM threat_controls")
    threat_control_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM risk_controls")
    risk_control_count = cur.fetchone()[0]

    print(f"Threats in database:           {threat_count}")
    print(f"Risks in database:             {risk_count}")
    print(f"Threat-control mappings:       {threat_control_count}")
    print(f"Risk-control mappings:         {risk_control_count}")

    # Sample data
    print("\nSample threats:")
    cur.execute("SELECT threat_id, name, category FROM threats ORDER BY threat_id LIMIT 5")
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]} [{row[2]}]")

    print("\nSample risks:")
    cur.execute("SELECT risk_id, title FROM risks WHERE risk_id IS NOT NULL ORDER BY risk_id LIMIT 5")
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]}")

    cur.close()

def main():
    print("=" * 80)
    print("GRC PLATFORM - THREAT & RISK COMPLETE IMPORT")
    print("=" * 80)
    print(f"\nExcel file: {EXCEL_PATH}")
    print(f"Database: {DB_URL}")

    # Verify Excel file exists
    if not os.path.exists(EXCEL_PATH):
        print(f"\n✗ ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print("\nLoading Excel workbook (this may take a moment)...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Available sheets: {wb.sheetnames}")

    print("\nConnecting to database...")
    conn = psycopg2.connect(DB_URL)

    # Step 1: Import catalogs first (threats and risks must exist before mappings)
    threat_count = import_threat_catalog(conn, wb)

    # Reload workbook for each operation (read_only mode limitation)
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    risk_count = import_risk_catalog(conn, wb)

    # Step 2: Import control mappings
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    risk_mapping_count = import_risk_control_mappings(conn, wb)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    threat_mapping_count = import_threat_control_mappings(conn, wb)

    # Step 3: Verify
    verify_import(conn)

    conn.close()

    print("\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    print(f"Threats imported:              {threat_count}")
    print(f"Risks imported:                {risk_count}")
    print(f"Risk-control mappings:         {risk_mapping_count}")
    print(f"Threat-control mappings:       {threat_mapping_count}")
    print("\n✓ Complete threat & risk import successful!")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n✗ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
