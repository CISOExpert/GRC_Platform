import os
from typing import Any, Dict, List, Set
import openpyxl
import psycopg2

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "reference_material", "secure-controls-framework-scf-2025-3-1 (1).xlsx")
SCF_SHEET_NAME = "SCF 2025.3.1"

# Columns: C = SCF # (control_id), mapping columns = AC (29) to NJ (274)
SCF_ID_COL = 3
MAPPING_COL_START = 29
MAPPING_COL_END = 274


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    value = value.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    value = " ".join(value.split())
    return value.strip()


def parse_refs(cell_value: Any) -> List[str]:
    """Split a cell value into individual external control refs."""
    if not cell_value:
        return []
    if isinstance(cell_value, str):
        # Split on newlines, semicolons, and (optionally) commas
        refs = []
        for part in cell_value.replace("\r", "").split("\n"):
            for subpart in part.split(";"):
                for ref in subpart.split(","):
                    ref = ref.strip()
                    if ref:
                        refs.append(ref)
        return refs
    return [str(cell_value).strip()]


def get_db_connection():
    return psycopg2.connect(
        dbname=os.getenv("SUPABASE_DB_NAME", "postgres"),
        user=os.getenv("SUPABASE_DB_USER", "postgres"),
        password=os.getenv("SUPABASE_DB_PASSWORD", "postgres"),
        host=os.getenv("SUPABASE_DB_HOST", "127.0.0.1"),
        port=int(os.getenv("SUPABASE_DB_PORT", "54322")),
    )


def load_scf_controls(conn) -> Dict[str, str]:
    with conn.cursor() as cur:
        cur.execute("SELECT id, control_id FROM scf_controls;")
        rows = cur.fetchall()
    return {control_id: scf_id for scf_id, control_id in rows}


def load_frameworks_by_header(conn) -> Dict[str, Dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, code, name, version, mapping_column_header
            FROM frameworks
            WHERE mapping_column_header IS NOT NULL
            """
        )
        rows = cur.fetchall()
    result: Dict[str, Dict[str, Any]] = {}
    for framework_id, code, name, version, header in rows:
        if header is None:
            continue
        key = normalize_header(header)
        result[key] = {
            "id": framework_id,
            "code": code,
            "name": name,
            "version": version,
            "mapping_column_header": key,
        }
    return result


def main():
    print("Connecting to database...")
    conn = get_db_connection()
    conn.autocommit = False
    try:
        print("Deleting all scf_control_mappings and external_controls...")
        with conn.cursor() as cur:
            cur.execute("DELETE FROM scf_control_mappings;")
            cur.execute("DELETE FROM external_controls;")
        print("All mappings and external controls deleted.")

        print("Loading SCF controls...")
        scf_controls = load_scf_controls(conn)
        print(f"Loaded {len(scf_controls)} SCF controls.")

        print("Loading frameworks by mapping_column_header...")
        frameworks = load_frameworks_by_header(conn)
        print(f"Loaded {len(frameworks)} frameworks with mapping_column_header.")

        print("Loading Excel workbook...")
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        sheet = wb[SCF_SHEET_NAME]

        # Map Excel column index to framework_id
        col_to_framework: Dict[int, str] = {}
        for col_idx in range(MAPPING_COL_START, MAPPING_COL_END):
            cell = sheet.cell(row=1, column=col_idx)
            norm = normalize_header(cell.value)
            fw = frameworks.get(norm)
            if fw:
                col_to_framework[col_idx] = fw["id"]

        print(f"Found {len(col_to_framework)} mapping columns with framework linkage.")

        total_mappings = 0
        total_controls = 0
        external_control_cache: Dict[tuple, str] = {}
        with conn.cursor() as cur:
            for row_idx in range(2, sheet.max_row + 1):
                if row_idx % 100 == 0:
                    print(f"Processed {row_idx} SCF rows...")
                scf_id_cell = sheet.cell(row=row_idx, column=SCF_ID_COL)
                scf_id = (scf_id_cell.value or "").strip() if isinstance(scf_id_cell.value, str) else scf_id_cell.value
                if not scf_id:
                    continue
                scf_db_id = scf_controls.get(scf_id)
                if not scf_db_id:
                    continue
                for col_idx, framework_id in col_to_framework.items():
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    refs = parse_refs(cell.value)
                    for ref in refs:
                        cache_key = (framework_id, ref)
                        external_id = external_control_cache.get(cache_key)
                        if not external_id:
                            cur.execute(
                                """
                                INSERT INTO external_controls (framework_id, ref_code, description)
                                VALUES (%s, %s, %s)
                                ON CONFLICT (framework_id, ref_code) DO NOTHING
                                RETURNING id
                                """,
                                (framework_id, ref, f"External control {ref}"),
                            )
                            ext_row = cur.fetchone()
                            if ext_row:
                                external_id = ext_row[0]
                            else:
                                cur.execute(
                                    "SELECT id FROM external_controls WHERE framework_id = %s AND ref_code = %s;",
                                    (framework_id, ref),
                                )
                                external_id = cur.fetchone()[0]
                            external_control_cache[cache_key] = external_id
                        cur.execute(
                            """
                            INSERT INTO scf_control_mappings (scf_control_id, external_control_id, framework_id)
                            VALUES (%s, %s, %s)
                            ON CONFLICT (scf_control_id, external_control_id, framework_id) DO NOTHING
                            """,
                            (scf_db_id, external_id, framework_id),
                        )
                        total_mappings += 1
                total_controls += 1
        wb.close()
        conn.commit()
        print(f"Rebuild complete. Imported {total_mappings} mappings for {total_controls} SCF controls.")
    except Exception as exc:
        print("Error during rebuild, rolling back:", exc)
        conn.rollback()
        raise
    finally:
        conn.close()

if __name__ == "__main__":
    main()
