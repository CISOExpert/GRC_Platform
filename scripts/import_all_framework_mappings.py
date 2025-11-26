#!/usr/bin/env python3
"""
Import all SCF control mappings from the Excel file to the database.
Handles multiple frameworks with hierarchy tracking.
"""

import openpyxl
import psycopg2
import psycopg2.extras
import json
import sys
import re
from typing import Dict, List, Tuple, Optional

# Database connection
DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 54322,
    'database': 'postgres',
    'user': 'postgres',
    'password': 'postgres'
}

# Major framework definitions
# Format: 'Framework Name': (column_index, code, version, name)
FRAMEWORKS = {
    # Already imported
    # 'NIST CSF v2.0': (93, 'NIST-CSF', '2.0', 'NIST Cybersecurity Framework'),
    
    # SOC 2 / AICPA
    'SOC 2': (29, 'SOC2-TSC', '2017-2022', 'SOC 2 Trust Service Criteria'),
    
    # CIS Controls
    'CIS v8.1': (31, 'CIS-CSC', 'v8.1', 'CIS Critical Security Controls'),
    'CIS v8.1 IG1': (32, 'CIS-CSC-IG1', 'v8.1', 'CIS Controls Implementation Group 1'),
    'CIS v8.1 IG2': (33, 'CIS-CSC-IG2', 'v8.1', 'CIS Controls Implementation Group 2'),
    'CIS v8.1 IG3': (34, 'CIS-CSC-IG3', 'v8.1', 'CIS Controls Implementation Group 3'),
    
    # COBIT
    'COBIT 2019': (35, 'COBIT', '2019', 'COBIT 2019 Framework'),
    
    # ISO Standards
    'ISO 27001:2022': (46, 'ISO-27001', '2022', 'ISO/IEC 27001:2022'),
    'ISO 27002:2022': (48, 'ISO-27002', '2022', 'ISO/IEC 27002:2022'),
    'ISO 27001:2013': (45, 'ISO-27001', '2013', 'ISO/IEC 27001:2013'),
    'ISO 27002:2013': (47, 'ISO-27002', '2013', 'ISO/IEC 27002:2013'),
    'ISO 27017:2015': (49, 'ISO-27017', '2015', 'ISO/IEC 27017:2015 Cloud Security'),
    'ISO 27018:2014': (50, 'ISO-27018', '2014', 'ISO/IEC 27018:2014 Cloud Privacy'),
    'ISO 27701:2019': (51, 'ISO-27701', '2019', 'ISO/IEC 27701:2019 Privacy'),
    'ISO 42001:2023': (55, 'ISO-42001', '2023', 'ISO/IEC 42001:2023 AI Management'),
    
    # NIST Standards
    'NIST 800-53 rev5': (68, 'NIST-800-53', 'rev5', 'NIST SP 800-53 Revision 5'),
    'NIST 800-53 rev4': (64, 'NIST-800-53', 'rev4', 'NIST SP 800-53 Revision 4'),
    'NIST 800-171 rev3': (87, 'NIST-800-171', 'rev3', 'NIST SP 800-171 Revision 3'),
    'NIST 800-171 rev2': (85, 'NIST-800-171', 'rev2', 'NIST SP 800-171 Revision 2'),
    'NIST 800-172': (89, 'NIST-800-172', '1.0', 'NIST SP 800-172 Enhanced Security'),
    'NIST CSF v1.1': (92, 'NIST-CSF', 'v1.1', 'NIST Cybersecurity Framework v1.1'),
    'NIST Privacy Framework': (61, 'NIST-PF', 'v1.0', 'NIST Privacy Framework'),
    'NIST AI RMF': (59, 'NIST-AI-RMF', 'v1.0', 'NIST AI Risk Management Framework'),
    
    # PCI DSS
    'PCI DSS v4.0.1': (96, 'PCI-DSS', 'v4.0.1', 'PCI Data Security Standard v4.0.1'),
    'PCI DSS v3.2': (95, 'PCI-DSS', 'v3.2', 'PCI Data Security Standard v3.2'),
    
    # HIPAA
    'HIPAA Security': (153, 'HIPAA', '2013', 'HIPAA Security Rule'),
    
    # GDPR
    'GDPR': (199, 'GDPR', '2016', 'EU General Data Protection Regulation'),
    
    # FedRAMP
    'FedRAMP Rev 5 High': (145, 'FedRAMP', 'R5-High', 'FedRAMP Rev 5 High Baseline'),
    'FedRAMP Rev 5 Moderate': (144, 'FedRAMP', 'R5-Moderate', 'FedRAMP Rev 5 Moderate Baseline'),
    'FedRAMP Rev 5 Low': (143, 'FedRAMP', 'R5-Low', 'FedRAMP Rev 5 Low Baseline'),
    
    # CMMC
    'CMMC 2.0': (80, 'CMMC', '2.0', 'Cybersecurity Maturity Model Certification'),
    
    # CSA CCM
    'CSA CCM v4': (37, 'CSA-CCM', 'v4', 'Cloud Security Alliance Cloud Controls Matrix'),
    
    # SOX
    'SOX': (164, 'SOX', '2002', 'Sarbanes-Oxley Act'),
    
    # State Laws
    'CCPA': (172, 'CCPA', '2022', 'California Consumer Privacy Act'),
    'NY DFS': (180, 'NYDFS', '2023', 'NY DFS 23 NYCRR 500'),
}

def connect_db():
    """Create database connection with autocommit."""
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = True
    return conn

def parse_control_reference(ref_code: str) -> Dict[str, str]:
    """
    Parse a control reference to determine its hierarchy level.
    Handles various formats from different frameworks.
    """
    ref_code = ref_code.strip()
    
    # Common patterns
    patterns = [
        # NIST CSF: GV.OC-01, PR.AA-02
        (r'^([A-Z]+)\.([A-Z]+)-(\d+)$', 'control', lambda m: {
            'function': m.group(1),
            'subcategory': f"{m.group(1)}.{m.group(2)}",
            'control_number': m.group(3)
        }),
        # NIST CSF Subcategory: GV.OC, PR.AA
        (r'^([A-Z]+)\.([A-Z]+)$', 'subcategory', lambda m: {
            'function': m.group(1),
            'subcategory': m.group(0),
            'control_number': None
        }),
        # NIST CSF Category: GV, PR, ID
        (r'^([A-Z]+)$', 'category', lambda m: {
            'function': m.group(1),
            'subcategory': None,
            'control_number': None
        }),
        # NIST 800-53: AC-1, SI-2(1)
        (r'^([A-Z]+)-(\d+)(?:\((\d+)\))?$', 'control', lambda m: {
            'family': m.group(1),
            'control_number': m.group(2),
            'enhancement': m.group(3)
        }),
        # ISO: A.5.1, A.8.2.3
        (r'^A\.(\d+(?:\.\d+)*)$', 'control', lambda m: {
            'control_number': m.group(1)
        }),
        # PCI: 1.1.1, 12.3.4
        (r'^(\d+(?:\.\d+)*)$', 'control', lambda m: {
            'control_number': m.group(1)
        }),
        # CIS: 1.1, 5.3
        (r'^(\d+)\.(\d+)$', 'control', lambda m: {
            'control_group': m.group(1),
            'control_number': m.group(2)
        }),
    ]
    
    for pattern, level, extractor in patterns:
        match = re.match(pattern, ref_code)
        if match:
            result = {
                'level': level,
                'full_ref': ref_code
            }
            result.update(extractor(match))
            return result
    
    # Generic fallback
    return {
        'level': 'control',
        'full_ref': ref_code
    }

def get_or_create_framework(cur, code: str, version: str, name: str) -> str:
    """Get existing framework or create new one."""
    # Check if exists
    cur.execute("""
        SELECT id FROM frameworks 
        WHERE code = %s AND version = %s;
    """, (code, version))
    
    result = cur.fetchone()
    if result:
        return result[0]
    
    # Create new framework
    cur.execute("""
        INSERT INTO frameworks (code, name, version)
        VALUES (%s, %s, %s)
        RETURNING id;
    """, (code, name, version))
    
    return cur.fetchone()[0]

def get_or_create_external_control(cur, framework_id: str, ref_code: str, hierarchy: Dict) -> str:
    """Get or create an external control entry (including categories/subcategories)."""
    # Check if exists
    cur.execute("""
        SELECT id FROM external_controls
        WHERE framework_id = %s AND ref_code = %s;
    """, (framework_id, ref_code))
    
    result = cur.fetchone()
    if result:
        return result[0]
    
    # Create description based on level
    level = hierarchy['level']
    if level == 'category':
        description = f"Category: {ref_code}"
    elif level == 'subcategory':
        description = f"Subcategory: {ref_code}"
    else:
        description = f"Control: {ref_code}"
    
    # Create metadata
    metadata = {
        'hierarchy_level': level,
        'is_category': level in ['category', 'subcategory']
    }
    metadata.update({k: v for k, v in hierarchy.items() if k not in ['level', 'full_ref'] and v is not None})
    
    # Insert
    cur.execute("""
        INSERT INTO external_controls (framework_id, ref_code, description, metadata)
        VALUES (%s, %s, %s, %s::jsonb)
        RETURNING id;
    """, (framework_id, ref_code, description, json.dumps(metadata)))
    
    return cur.fetchone()[0]

def determine_mapping_strength(framework_name: str) -> str:
    """Determine mapping strength based on framework name."""
    if '(partial' in framework_name.lower():
        return 'partial'
    return 'exact'

def import_framework(ws, framework_display_name: str, column_index: int, 
                     framework_code: str, framework_version: str, framework_name: str):
    """Import mappings for a specific framework."""
    
    print(f"\n{'='*70}")
    print(f"Importing: {framework_display_name}")
    print(f"{'='*70}")
    
    conn = connect_db()
    cur = conn.cursor()
    
    try:
        # Get or create framework
        framework_id = get_or_create_framework(cur, framework_code, framework_version, framework_name)
        print(f"Framework ID: {framework_id}")
        
        # Check if already imported
        cur.execute("""
            SELECT COUNT(*) FROM scf_control_mappings
            WHERE framework_id = %s;
        """, (framework_id,))
        
        existing_count = cur.fetchone()[0]
        if existing_count > 0:
            print(f"⚠ Already has {existing_count} mappings, skipping...")
            return
        
        # Determine mapping strength
        default_strength = determine_mapping_strength(framework_display_name)
        
        # Cache for external control IDs
        external_control_cache = {}
        
        # Statistics
        stats = {
            'scf_controls_processed': 0,
            'controls_with_mappings': 0,
            'total_mappings_created': 0,
            'category_mappings': 0,
            'subcategory_mappings': 0,
            'control_mappings': 0,
            'errors': 0
        }
        
        print("Processing mappings...")
        
        # Process each SCF control
        for row_num in range(2, ws.max_row + 1):
            scf_control_ref = ws.cell(row_num, 3).value  # Column 3: SCF #
            framework_mappings = ws.cell(row_num, column_index).value
            
            if not scf_control_ref:
                continue
            
            stats['scf_controls_processed'] += 1
            
            # Get SCF control ID
            cur.execute("""
                SELECT id FROM scf_controls WHERE control_id = %s;
            """, (scf_control_ref,))
            
            scf_result = cur.fetchone()
            if not scf_result:
                continue
            
            scf_control_id = scf_result[0]
            
            # Skip if no mappings
            if not framework_mappings:
                continue
            
            stats['controls_with_mappings'] += 1
            
            # Parse mappings (split by newlines)
            mapping_refs = str(framework_mappings).strip().split('\n')
            
            for mapping_ref in mapping_refs:
                mapping_ref = mapping_ref.strip()
                if not mapping_ref:
                    continue
                
                try:
                    # Parse the reference
                    hierarchy = parse_control_reference(mapping_ref)
                    
                    # Get or create external control
                    if mapping_ref in external_control_cache:
                        external_control_id = external_control_cache[mapping_ref]
                    else:
                        external_control_id = get_or_create_external_control(
                            cur, framework_id, mapping_ref, hierarchy
                        )
                        external_control_cache[mapping_ref] = external_control_id
                    
                    # Create mapping notes
                    notes = f"Mapping Level: {hierarchy['level']}"
                    for key, value in hierarchy.items():
                        if key not in ['level', 'full_ref'] and value is not None:
                            notes += f" | {key}: {value}"
                    
                    # Check if mapping already exists
                    cur.execute("""
                        SELECT id FROM scf_control_mappings
                        WHERE scf_control_id = %s AND external_control_id = %s;
                    """, (scf_control_id, external_control_id))
                    
                    if cur.fetchone():
                        continue
                    
                    # Insert mapping
                    cur.execute("""
                        INSERT INTO scf_control_mappings 
                        (scf_control_id, external_control_id, framework_id, mapping_strength, confidence, notes)
                        VALUES (%s, %s, %s, %s, 95, %s);
                    """, (scf_control_id, external_control_id, framework_id, default_strength, notes))
                    
                    stats['total_mappings_created'] += 1
                    
                    # Track by level
                    if hierarchy['level'] == 'category':
                        stats['category_mappings'] += 1
                    elif hierarchy['level'] == 'subcategory':
                        stats['subcategory_mappings'] += 1
                    else:
                        stats['control_mappings'] += 1
                    
                except Exception as e:
                    stats['errors'] += 1
                    if stats['errors'] <= 5:  # Only print first 5 errors
                        print(f"  ERROR: {scf_control_ref} -> {mapping_ref}: {e}")
                    continue
            
            # Progress indicator
            if row_num % 500 == 0:
                print(f"  Processed {row_num - 1} rows...")
        
        # Print statistics
        print(f"\n{'Results:':40}")
        print(f"  SCF controls with mappings: {stats['controls_with_mappings']}")
        print(f"  Total mappings created: {stats['total_mappings_created']}")
        if stats['category_mappings'] > 0 or stats['subcategory_mappings'] > 0:
            print(f"    - Category-level: {stats['category_mappings']}")
            print(f"    - Subcategory-level: {stats['subcategory_mappings']}")
        print(f"    - Control-level: {stats['control_mappings']}")
        if stats['errors'] > 0:
            print(f"  Errors: {stats['errors']}")
        
        print(f"✓ Complete")
        
    except Exception as e:
        print(f"✗ FAILED: {e}")
    finally:
        cur.close()
        conn.close()

def main():
    """Import all frameworks."""
    excel_path = '/Users/doneil/SynologyDrive/GRC_Unified_Platform/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx'
    
    print("="*70)
    print("IMPORTING ALL FRAMEWORK MAPPINGS")
    print("="*70)
    print(f"\nLoading Excel file: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['SCF 2025.3.1']
    
    print(f"Found {len(FRAMEWORKS)} frameworks to import\n")
    
    # Import each framework
    for display_name, (col_idx, code, version, name) in FRAMEWORKS.items():
        import_framework(ws, display_name, col_idx, code, version, name)
    
    wb.close()
    
    print("\n" + "="*70)
    print("ALL IMPORTS COMPLETE")
    print("="*70)

if __name__ == '__main__':
    main()
