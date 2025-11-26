#!/usr/bin/env python3
"""
Import SCF control mappings from the Excel file to the database.
Handles both category-level and control-level mappings with hierarchy tracking.
"""

import openpyxl
import psycopg2
import psycopg2.extras
import json
import sys
import re
from typing import Dict, List, Tuple, Optional

# Database connection
DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 54322,
    'database': 'postgres',
    'user': 'postgres',
    'password': 'postgres'
}

# Framework column mappings in Excel
FRAMEWORK_COLUMNS = {
    'NIST CSF v2.0': 93,
    # Add more frameworks later
}

def connect_db():
    """Create database connection with autocommit."""
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = True
    return conn

def parse_control_reference(ref_code: str) -> Dict[str, str]:
    """
    Parse a control reference to determine its hierarchy level.
    
    Examples:
    - "GV" -> category (function level)
    - "GV.OC" -> subcategory
    - "GV.OC-01" -> control (specific)
    - "PR.AA" -> subcategory
    - "PR.AA-02" -> control
    
    Returns dict with: level, function, subcategory, control_number
    """
    ref_code = ref_code.strip()
    
    # Pattern: XX.XX-## (specific control)
    control_pattern = r'^([A-Z]+)\.([A-Z]+)-(\d+)$'
    # Pattern: XX.XX (subcategory)
    subcat_pattern = r'^([A-Z]+)\.([A-Z]+)$'
    # Pattern: XX (function/category)
    func_pattern = r'^([A-Z]+)$'
    
    control_match = re.match(control_pattern, ref_code)
    if control_match:
        return {
            'level': 'control',
            'function': control_match.group(1),
            'subcategory': f"{control_match.group(1)}.{control_match.group(2)}",
            'control_number': control_match.group(3),
            'full_ref': ref_code
        }
    
    subcat_match = re.match(subcat_pattern, ref_code)
    if subcat_match:
        return {
            'level': 'subcategory',
            'function': subcat_match.group(1),
            'subcategory': ref_code,
            'control_number': None,
            'full_ref': ref_code
        }
    
    func_match = re.match(func_pattern, ref_code)
    if func_match:
        return {
            'level': 'category',
            'function': ref_code,
            'subcategory': None,
            'control_number': None,
            'full_ref': ref_code
        }
    
    # Unknown format
    return {
        'level': 'unknown',
        'function': None,
        'subcategory': None,
        'control_number': None,
        'full_ref': ref_code
    }

def get_or_create_category_control(cur, framework_id: str, ref_code: str, hierarchy: Dict) -> str:
    """
    Get or create a category-level or subcategory-level control entry.
    These are higher-level groupings that don't have detailed descriptions.
    """
    # Check if it already exists
    cur.execute("""
        SELECT id FROM external_controls
        WHERE framework_id = %s AND ref_code = %s;
    """, (framework_id, ref_code))
    
    result = cur.fetchone()
    if result:
        return result[0]
    
    # Create description based on level
    if hierarchy['level'] == 'category':
        description = f"NIST CSF v2.0 Function: {ref_code}"
    elif hierarchy['level'] == 'subcategory':
        description = f"NIST CSF v2.0 Subcategory: {ref_code}"
    else:
        description = f"NIST CSF v2.0 Reference: {ref_code}"
    
    # Create metadata
    metadata = {
        'hierarchy_level': hierarchy['level'],
        'function': hierarchy['function'],
        'is_category': True
    }
    
    if hierarchy['subcategory']:
        metadata['subcategory'] = hierarchy['subcategory']
    
    # Insert
    cur.execute("""
        INSERT INTO external_controls (framework_id, ref_code, description, metadata)
        VALUES (%s, %s, %s, %s::jsonb)
        RETURNING id;
    """, (framework_id, ref_code, description, json.dumps(metadata)))
    
    return cur.fetchone()[0]

def determine_mapping_strength(framework_name: str) -> str:
    """
    Determine mapping strength based on framework name.
    If "(partial)" appears in name, use 'partial', otherwise 'exact'.
    """
    if '(partial' in framework_name.lower():
        return 'partial'
    return 'exact'

def import_framework_mappings(excel_path: str, framework_name: str, column_index: int):
    """Import mappings for a specific framework."""
    print(f"\n{'='*60}")
    print(f"Importing {framework_name} mappings")
    print(f"{'='*60}\n")
    
    # Load Excel file
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['SCF 2025.3.1']
    
    # Connect to database
    conn = connect_db()
    cur = conn.cursor()
    
    # Get framework ID
    cur.execute("""
        SELECT id FROM frameworks WHERE code = 'NIST-CSF' AND version = '2.0';
    """)
    framework_result = cur.fetchone()
    if not framework_result:
        print("ERROR: NIST CSF v2.0 framework not found in database!")
        return
    
    framework_id = framework_result[0]
    print(f"Framework ID: {framework_id}")
    
    # Determine default mapping strength
    default_strength = determine_mapping_strength(framework_name)
    print(f"Default mapping strength: {default_strength}")
    
    # Cache for external control IDs to avoid repeated lookups
    external_control_cache = {}
    
    # Statistics
    stats = {
        'total_scf_controls': 0,
        'controls_with_mappings': 0,
        'total_mappings_created': 0,
        'category_mappings': 0,
        'subcategory_mappings': 0,
        'control_mappings': 0,
        'skipped_unknown': 0,
        'errors': 0
    }
    
    print("\nProcessing mappings...")
    
    # Process each SCF control
    for row_num in range(2, ws.max_row + 1):
        # Column 3: SCF # (e.g., "GOV-01")
        scf_control_ref = ws.cell(row_num, 3).value
        # Column 93: NIST CSF v2.0 mappings
        nist_mappings = ws.cell(row_num, column_index).value
        
        if not scf_control_ref:
            continue
        
        stats['total_scf_controls'] += 1
        
        # Get SCF control ID from database
        cur.execute("""
            SELECT id FROM scf_controls WHERE control_id = %s;
        """, (scf_control_ref,))
        
        scf_result = cur.fetchone()
        if not scf_result:
            # Try alternate format (some might have spaces or different naming)
            continue
        
        scf_control_id = scf_result[0]
        
        # Skip if no mappings
        if not nist_mappings:
            continue
        
        stats['controls_with_mappings'] += 1
        
        # Parse mappings (split by newlines)
        mapping_refs = str(nist_mappings).strip().split('\n')
        
        for mapping_ref in mapping_refs:
            mapping_ref = mapping_ref.strip()
            if not mapping_ref:
                continue
            
            try:
                # Parse the reference to determine hierarchy
                hierarchy = parse_control_reference(mapping_ref)
                
                if hierarchy['level'] == 'unknown':
                    stats['skipped_unknown'] += 1
                    continue
                
                # Get or create external control ID
                if mapping_ref in external_control_cache:
                    external_control_id = external_control_cache[mapping_ref]
                else:
                    # Check if it's a specific control (should already exist)
                    cur.execute("""
                        SELECT id FROM external_controls
                        WHERE framework_id = %s AND ref_code = %s;
                    """, (framework_id, mapping_ref))
                    
                    result = cur.fetchone()
                    if result:
                        external_control_id = result[0]
                    else:
                        # Create category or subcategory entry
                        external_control_id = get_or_create_category_control(
                            cur, framework_id, mapping_ref, hierarchy
                        )
                    
                    external_control_cache[mapping_ref] = external_control_id
                
                # Create mapping notes with hierarchy info
                notes = f"Mapping Level: {hierarchy['level']}"
                if hierarchy['function']:
                    notes += f" | Function: {hierarchy['function']}"
                if hierarchy['subcategory']:
                    notes += f" | Subcategory: {hierarchy['subcategory']}"
                
                # Check if mapping already exists
                cur.execute("""
                    SELECT id FROM scf_control_mappings
                    WHERE scf_control_id = %s AND external_control_id = %s;
                """, (scf_control_id, external_control_id))
                
                if cur.fetchone():
                    continue  # Skip duplicates
                
                # Insert mapping
                cur.execute("""
                    INSERT INTO scf_control_mappings 
                    (scf_control_id, external_control_id, framework_id, mapping_strength, confidence, notes)
                    VALUES (%s, %s, %s, %s, 95, %s);
                """, (scf_control_id, external_control_id, framework_id, default_strength, notes))
                
                stats['total_mappings_created'] += 1
                
                # Track by level
                if hierarchy['level'] == 'category':
                    stats['category_mappings'] += 1
                elif hierarchy['level'] == 'subcategory':
                    stats['subcategory_mappings'] += 1
                elif hierarchy['level'] == 'control':
                    stats['control_mappings'] += 1
                
            except Exception as e:
                print(f"ERROR processing mapping {scf_control_ref} -> {mapping_ref}: {e}")
                stats['errors'] += 1
                continue
        
        # Progress indicator
        if row_num % 100 == 0:
            print(f"  Processed {row_num - 1} SCF controls...")
    
    # Print statistics
    print(f"\n{'='*60}")
    print("Import Statistics")
    print(f"{'='*60}")
    print(f"Total SCF controls processed: {stats['total_scf_controls']}")
    print(f"SCF controls with mappings: {stats['controls_with_mappings']}")
    print(f"Total mappings created: {stats['total_mappings_created']}")
    print(f"  - Category-level mappings: {stats['category_mappings']}")
    print(f"  - Subcategory-level mappings: {stats['subcategory_mappings']}")
    print(f"  - Control-level mappings: {stats['control_mappings']}")
    print(f"Skipped (unknown format): {stats['skipped_unknown']}")
    print(f"Errors: {stats['errors']}")
    
    # Verify mappings
    cur.execute("""
        SELECT COUNT(*) FROM scf_control_mappings
        WHERE framework_id = %s;
    """, (framework_id,))
    total_in_db = cur.fetchone()[0]
    print(f"\nTotal mappings in database for {framework_name}: {total_in_db}")
    
    cur.close()
    conn.close()
    wb.close()

if __name__ == '__main__':
    excel_path = '/Users/doneil/SynologyDrive/GRC_Unified_Platform/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx'
    
    # Import NIST CSF v2.0 mappings
    import_framework_mappings(excel_path, 'NIST CSF v2.0', FRAMEWORK_COLUMNS['NIST CSF v2.0'])
    
    print("\n✓ Import complete!")
