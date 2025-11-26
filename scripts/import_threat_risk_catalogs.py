#!/usr/bin/env python3
"""
Import Threat and Risk Catalogs from SCF Excel
Imports threats (NT-*, MT-*) and risks (R-*) with their groupings and metadata
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import sys
import re

DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = "/Users/doneil/SynologyDrive/GRC_Unified_Platform/reference_material/secure-controls-framework-scf-2025-3-1 (1).xlsx"

def clean_text(text):
    if text is None:
        return None
    text = str(text).strip()
    text = re.sub(r'\n+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text if text else None

def import_threats(conn):
    """Import threats from Threat Catalog tab"""
    print("\n" + "=" * 80)
    print("IMPORTING THREATS")
    print("=" * 80)
    
    cur = conn.cursor()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['Threat Catalog']
    
    threats = []
    current_grouping = None
    
    # Start from row 8 (after headers)
    for row in ws.iter_rows(min_row=8, values_only=True):
        grouping = clean_text(row[0])
        threat_id = clean_text(row[1])
        name = clean_text(row[2])
        description = clean_text(row[3])
        
        # Track current grouping
        if grouping:
            current_grouping = grouping
        
        # Only process rows with threat IDs
        if threat_id and (threat_id.startswith('NT-') or threat_id.startswith('MT-') or threat_id.startswith('FT-')):
            threats.append({
                'threat_id': threat_id,
                'name': name,
                'threat_grouping': current_grouping,
                'description': description,
                'category': 'natural' if threat_id.startswith('NT-') else 'manmade'
            })
    
    print(f"Extracted {len(threats)} threats")
    
    # Insert threats
    for threat in threats:
        cur.execute("""
            INSERT INTO threats (threat_id, name, threat_grouping, description, category)
            VALUES (%s, %s, %s, %s, %s::threat_category)
            ON CONFLICT (threat_id) DO UPDATE
            SET name = EXCLUDED.name,
                threat_grouping = EXCLUDED.threat_grouping,
                description = EXCLUDED.description,
                category = EXCLUDED.category
        """, (
            threat['threat_id'],
            threat['name'],
            threat['threat_grouping'],
            threat['description'],
            threat['category']
        ))
    
    conn.commit()
    print(f"✓ Imported {len(threats)} threats")
    cur.close()
    return len(threats)

def import_risks(conn):
    """Import risks from Risk Catalog tab"""
    print("\n" + "=" * 80)
    print("IMPORTING RISKS")
    print("=" * 80)
    
    cur = conn.cursor()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['Risk Catalog']
    
    risks = []
    current_grouping = None
    
    # Start from row 8 (after headers)
    for row in ws.iter_rows(min_row=8, values_only=True):
        grouping = clean_text(row[0])
        risk_id = clean_text(row[1])
        name = clean_text(row[2])
        description = clean_text(row[3])
        nist_csf_function = clean_text(row[4]) if len(row) > 4 else None
        
        # Track current grouping
        if grouping:
            current_grouping = grouping
        
        # Only process rows with risk IDs
        if risk_id and risk_id.startswith('R-'):
            risks.append({
                'risk_id': risk_id,
                'title': name,
                'risk_grouping': current_grouping,
                'description': description,
                'nist_csf_function': nist_csf_function
            })
    
    print(f"Extracted {len(risks)} risks")
    
    # Insert risks (no org_id for catalog risks)
    for risk in risks:
        cur.execute("""
            INSERT INTO risks (risk_id, title, risk_grouping, description, nist_csf_function, status)
            VALUES (%s, %s, %s, %s, %s, 'catalog')
            ON CONFLICT (risk_id) DO UPDATE
            SET title = EXCLUDED.title,
                risk_grouping = EXCLUDED.risk_grouping,
                description = EXCLUDED.description,
                nist_csf_function = EXCLUDED.nist_csf_function
        """, (
            risk['risk_id'],
            risk['title'],
            risk['risk_grouping'],
            risk['description'],
            risk['nist_csf_function']
        ))
    
    conn.commit()
    print(f"✓ Imported {len(risks)} risks")
    cur.close()
    return len(risks)

def main():
    print("=" * 80)
    print("SCF THREAT & RISK CATALOG IMPORT")
    print("=" * 80)
    
    conn = psycopg2.connect(DB_URL)
    
    threat_count = import_threats(conn)
    risk_count = import_risks(conn)
    
    conn.close()
    
    print("\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    print(f"Total threats imported: {threat_count}")
    print(f"Total risks imported:   {risk_count}")
    print("\n✓ Threat & Risk catalog import successful!")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n✗ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
